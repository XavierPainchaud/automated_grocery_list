import os
from openpyxl import load_workbook

def update_weekly_items(items):
    """
    Mettre à jour l'inventaire en fonction des ingrédients dans le fichier texte
    """
    try:
        # Chemin du fichier Excel d'inventaire
        inventory_file = r'C:\github_repo\automated_grocery_list\grocery\grocery.xlsx'
        
        # Vérifier si le fichier Excel existe
        if not os.path.exists(inventory_file):
            print(f"Erreur : Le fichier Excel {inventory_file} n'existe pas.")
            return

        # Charger l'inventaire
        workbook = load_workbook(inventory_file)
        if "inventory" not in workbook.sheetnames:
            print("Erreur : La feuille 'inventory' n'existe pas dans le fichier Excel.")
            return
        
        sheet = workbook["inventory"]
        
        # Identifier les colonnes pertinentes
        headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}
        if "Item" not in headers or "In stock" not in headers:
            print("Les colonnes 'Item' ou 'In stock' sont absentes du fichier Excel.")
            return
        
        item_col = headers["Item"]
        in_stock_col = headers["In stock"]


        # Mettre à jour la colonne 'In stock' à 'Yes' pour les ingrédients présents
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
            item_cell = row[item_col - 1]  # Colonne 'Item'
            in_stock_cell = row[in_stock_col - 1]  # Colonne 'In stock'
            if item_cell.value in items:
                answer = input(f"Avez vous besoin de {item_cell.value} (Y or N)").strip().upper()
                while answer not in ['Y', 'N']:
                    print("Choix invalide")
                    answer = input(f"Avez vous besoin de {item_cell.value} (Y or N)").strip().upper()
                if answer == 'Y' and in_stock_cell.value == "Yes":
                    in_stock_cell.value = "No"
                    print(f'Besoin de {item_cell.value} cette semaine') 
                elif answer == 'N': 
                    print(f'Pas besoin de {item_cell.value} cette semaine.')
                else:
                    print(f"{item_cell.value} déjà défini comme requis dans l'inventaire.") 
        
        # Sauvegarder les changements dans le fichier Excel
        workbook.save(inventory_file)
        print("L'inventaire a été mis à jour avec succès.")
    
    except Exception as e:
        print(f"Erreur lors de la mise à jour de l'inventaire : {str(e)}")

if __name__ == "__main__":
    items = ["Banane", "Brocoli", "Fruits", "Lait", "Oeuf", "Pain", "Pâte", "Poivron", "Yogourt grec"]
    update_weekly_items(items)
