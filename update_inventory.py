import os
from openpyxl import load_workbook

def read_shopping_list(file_path):
    """
    Lire les ingrédients depuis le fichier texte.
    """
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            lines = file.readlines()
        shopping_list = [line.strip() for line in lines if line.strip()]  # Nettoyer les lignes
        return shopping_list
    except Exception as e:
        print(f"Erreur lors de la lecture du fichier texte : {str(e)}")
        return []

def update_inventory_from_text(file_path):
    """
    Mettre à jour l'inventaire en fonction des ingrédients dans le fichier texte
    """
    try:
        # Chemin du fichier Excel d'inventaire
        inventory_file = r'C:\github_repo\automated_grocery_list\grocery\grocery.xlsx'
        
        # Vérifier si le fichier Excel existe
        if not os.path.exists(inventory_file):
            print(f"Erreur : Le fichier Excel {inventory_file} n'existe pas.")
            return

        # Charger l'inventaire
        workbook = load_workbook(inventory_file)
        if "inventory" not in workbook.sheetnames:
            print("Erreur : La feuille 'inventory' n'existe pas dans le fichier Excel.")
            return
        
        sheet = workbook["inventory"]
        
        # Lire la liste d'épicerie depuis le fichier texte
        shopping_list = read_shopping_list(file_path)
        
        if not shopping_list:
            print("La liste d'épicerie est vide ou invalide.")
            return
        
        # Identifier les colonnes pertinentes
        headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}
        if "Item" not in headers or "In stock" not in headers:
            print("Les colonnes 'Item' ou 'In stock' sont absentes du fichier Excel.")
            return
        
        item_col = headers["Item"]
        in_stock_col = headers["In stock"]

        # Mettre à jour la colonne 'In stock' à 'Yes' pour les ingrédients présents
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
            item_cell = row[item_col - 1]  # Colonne 'Item'
            in_stock_cell = row[in_stock_col - 1]  # Colonne 'In stock'
            if item_cell.value in shopping_list:
                in_stock_cell.value = "Yes"
        
        # Sauvegarder les changements dans le fichier Excel
        workbook.save(inventory_file)
        print("L'inventaire a été mis à jour avec succès.")
    
    except Exception as e:
        print(f"Erreur lors de la mise à jour de l'inventaire : {str(e)}")

if __name__ == "__main__":
    # Chemin du fichier texte
    shopping_list_path = r"C:\Users\Xavier\Downloads\liste_epicerie.txt"
    
    if not os.path.exists(shopping_list_path):
        print(f"Erreur : Le fichier {shopping_list_path} n'existe pas.")
    else:
        update_inventory_from_text(shopping_list_path)
